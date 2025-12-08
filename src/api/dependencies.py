"""
FastAPI dependencies for the Taxdown API.

Provides dependency injection for database connections, services, and authentication.
"""

from fastapi import Depends, HTTPException, Security
from fastapi.security import APIKeyHeader
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from typing import Generator, Optional

from src.api.config import get_settings, APISettings
from src.services import (
    AssessmentAnalyzer,
    AppealGenerator,
    GeneratorConfig,
    PDFGenerator,
    PortfolioService,
    BulkAnalysisService,
    PortfolioAnalytics,
)

# Database engine (singleton)
_engine = None


def get_engine():
    """Get database engine singleton."""
    global _engine
    if _engine is None:
        settings = get_settings()
        _engine = create_engine(
            settings.database_url,
            pool_size=settings.database_pool_size,
            max_overflow=settings.database_max_overflow,
        )
    return _engine


def get_db() -> Generator:
    """Database session dependency."""
    engine = get_engine()
    with engine.connect() as connection:
        yield connection


# API Key Security
api_key_header = APIKeyHeader(name="X-API-Key", auto_error=False)


async def verify_api_key(
    api_key: Optional[str] = Security(api_key_header),
    settings: APISettings = Depends(get_settings),
) -> Optional[str]:
    """Verify API key if required."""
    if not settings.require_api_key:
        return None

    if not api_key:
        raise HTTPException(status_code=401, detail="API key required")

    if api_key not in settings.api_keys:
        raise HTTPException(status_code=403, detail="Invalid API key")

    return api_key


# ============================================================================
# SERVICE DEPENDENCIES
# ============================================================================


def get_assessment_analyzer() -> AssessmentAnalyzer:
    """Get AssessmentAnalyzer instance."""
    return AssessmentAnalyzer(get_engine())


def get_appeal_generator() -> AppealGenerator:
    """Get AppealGenerator instance with default config."""
    return AppealGenerator(get_engine())


def get_appeal_generator_with_config(config: GeneratorConfig) -> AppealGenerator:
    """Get AppealGenerator instance with custom config."""
    return AppealGenerator(get_engine(), config)


def get_pdf_generator() -> PDFGenerator:
    """Get PDFGenerator instance."""
    return PDFGenerator()


def get_portfolio_service() -> PortfolioService:
    """Get PortfolioService instance."""
    return PortfolioService(get_engine())


def get_bulk_analysis_service() -> BulkAnalysisService:
    """Get BulkAnalysisService instance."""
    return BulkAnalysisService(get_engine())


def get_portfolio_analytics() -> PortfolioAnalytics:
    """Get PortfolioAnalytics instance."""
    return PortfolioAnalytics(get_engine())


def get_report_generator():
    """
    Get PortfolioReportGenerator instance.

    Returns a simple implementation that uses PortfolioService for data.
    """
    from sqlalchemy import text
    import csv
    import json

    class SimpleReportGenerator:
        """Simple report generator using portfolio service."""

        def __init__(self, engine):
            self.engine = engine
            self.portfolio_service = PortfolioService(engine)

        def generate_pdf_report(self, portfolio_id: str, output_path: str, options: dict = None):
            """Generate PDF report (requires reportlab)."""
            # Get portfolio data
            data = self.get_portfolio_data(portfolio_id)

            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet

                doc = SimpleDocTemplate(output_path, pagesize=letter)
                elements = []
                styles = getSampleStyleSheet()

                # Title
                elements.append(Paragraph(f"Portfolio Report: {data['name']}", styles['Title']))
                elements.append(Spacer(1, 20))

                # Summary
                elements.append(Paragraph("Summary", styles['Heading2']))
                summary_data = [
                    ["Total Properties", str(data.get('total_properties', 0))],
                    ["Total Market Value", f"${data.get('total_market_value', 0):,.2f}"],
                    ["Total Assessed Value", f"${data.get('total_assessed_value', 0):,.2f}"],
                    ["Total Potential Savings", f"${data.get('total_potential_savings', 0):,.2f}"],
                    ["Appeal Candidates", str(data.get('appeal_candidates', 0))],
                ]
                summary_table = Table(summary_data, colWidths=[200, 200])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                elements.append(summary_table)
                elements.append(Spacer(1, 20))

                # Properties
                if data.get('properties'):
                    elements.append(Paragraph("Properties", styles['Heading2']))
                    prop_data = [["Parcel ID", "Address", "Value", "Score", "Action"]]
                    for prop in data['properties'][:50]:  # Limit to 50
                        prop_data.append([
                            prop.get('parcel_id', ''),
                            prop.get('address', '')[:30] if prop.get('address') else '',
                            f"${prop.get('market_value', 0):,.0f}",
                            str(prop.get('fairness_score', 'N/A')),
                            prop.get('recommended_action', 'N/A'),
                        ])
                    prop_table = Table(prop_data, colWidths=[80, 150, 80, 50, 60])
                    prop_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    elements.append(prop_table)

                doc.build(elements)

            except ImportError:
                # Fallback: generate simple text file
                with open(output_path, 'w') as f:
                    f.write(f"Portfolio Report: {data['name']}\n")
                    f.write("=" * 50 + "\n\n")
                    f.write(f"Total Properties: {data.get('total_properties', 0)}\n")
                    f.write(f"Total Market Value: ${data.get('total_market_value', 0):,.2f}\n")
                    f.write(f"Appeal Candidates: {data.get('appeal_candidates', 0)}\n")

        def generate_csv_export(self, portfolio_id: str, output_path: str, include_analysis: bool = True):
            """Generate CSV export of portfolio properties."""
            data = self.get_portfolio_data(portfolio_id)

            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                # Header
                headers = [
                    "Parcel ID", "Address", "City", "Owner",
                    "Market Value", "Assessed Value", "Ownership Type",
                ]
                if include_analysis:
                    headers.extend(["Fairness Score", "Recommendation", "Potential Savings"])

                writer.writerow(headers)

                # Data rows
                for prop in data.get('properties', []):
                    row = [
                        prop.get('parcel_id', ''),
                        prop.get('address', ''),
                        prop.get('city', ''),
                        prop.get('owner_name', ''),
                        prop.get('market_value', 0),
                        prop.get('assessed_value', 0),
                        prop.get('ownership_type', ''),
                    ]
                    if include_analysis:
                        row.extend([
                            prop.get('fairness_score', ''),
                            prop.get('recommended_action', ''),
                            prop.get('estimated_savings', 0),
                        ])
                    writer.writerow(row)

        def generate_excel_export(self, portfolio_id: str, output_path: str):
            """Generate Excel export (requires openpyxl)."""
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill

                data = self.get_portfolio_data(portfolio_id)
                wb = Workbook()

                # Summary sheet
                ws_summary = wb.active
                ws_summary.title = "Summary"
                ws_summary['A1'] = "Portfolio Summary"
                ws_summary['A1'].font = Font(bold=True, size=14)
                ws_summary['A3'] = "Name"
                ws_summary['B3'] = data.get('name', '')
                ws_summary['A4'] = "Total Properties"
                ws_summary['B4'] = data.get('total_properties', 0)
                ws_summary['A5'] = "Total Market Value"
                ws_summary['B5'] = data.get('total_market_value', 0)
                ws_summary['A6'] = "Appeal Candidates"
                ws_summary['B6'] = data.get('appeal_candidates', 0)

                # Properties sheet
                ws_props = wb.create_sheet("Properties")
                headers = ["Parcel ID", "Address", "City", "Market Value",
                          "Assessed Value", "Fairness Score", "Recommendation"]
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for col, header in enumerate(headers, 1):
                    cell = ws_props.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                for row_num, prop in enumerate(data.get('properties', []), 2):
                    ws_props.cell(row=row_num, column=1, value=prop.get('parcel_id', ''))
                    ws_props.cell(row=row_num, column=2, value=prop.get('address', ''))
                    ws_props.cell(row=row_num, column=3, value=prop.get('city', ''))
                    ws_props.cell(row=row_num, column=4, value=prop.get('market_value', 0))
                    ws_props.cell(row=row_num, column=5, value=prop.get('assessed_value', 0))
                    ws_props.cell(row=row_num, column=6, value=prop.get('fairness_score', ''))
                    ws_props.cell(row=row_num, column=7, value=prop.get('recommended_action', ''))

                wb.save(output_path)

            except ImportError:
                # Fallback to CSV
                self.generate_csv_export(portfolio_id, output_path.replace('.xlsx', '.csv'), True)

        def generate_summary_text(self, portfolio_id: str) -> str:
            """Generate plain text summary."""
            data = self.get_portfolio_data(portfolio_id)

            lines = [
                f"Portfolio: {data.get('name', 'Unknown')}",
                f"Description: {data.get('description', 'N/A')}",
                "",
                "Summary:",
                f"  Total Properties: {data.get('total_properties', 0)}",
                f"  Total Market Value: ${data.get('total_market_value', 0):,.2f}",
                f"  Total Assessed Value: ${data.get('total_assessed_value', 0):,.2f}",
                f"  Estimated Annual Tax: ${data.get('estimated_annual_tax', 0):,.2f}",
                f"  Total Potential Savings: ${data.get('total_potential_savings', 0):,.2f}",
                f"  Appeal Candidates: {data.get('appeal_candidates', 0)}",
            ]

            return "\n".join(lines)

        def get_portfolio_data(self, portfolio_id: str) -> dict:
            """Get portfolio data for report generation."""
            portfolio = self.portfolio_service.get_portfolio(portfolio_id)
            if not portfolio:
                raise ValueError(f"Portfolio {portfolio_id} not found")

            properties = self.portfolio_service.get_portfolio_properties(portfolio_id)

            def cents_to_dollars(cents):
                return cents / 100.0 if cents else 0

            return {
                "id": str(portfolio.id),
                "name": portfolio.name,
                "description": portfolio.description,
                "total_properties": len(properties),
                "total_market_value": cents_to_dollars(portfolio.total_market_value_cents),
                "total_assessed_value": cents_to_dollars(portfolio.total_assessed_value_cents),
                "estimated_annual_tax": cents_to_dollars(portfolio.estimated_annual_tax_cents),
                "total_potential_savings": cents_to_dollars(portfolio.total_potential_savings_cents),
                "appeal_candidates": portfolio.appeal_candidates,
                "properties": [
                    {
                        "parcel_id": p.parcel_id,
                        "address": p.address,
                        "city": p.city,
                        "owner_name": p.owner_name,
                        "market_value": cents_to_dollars(p.market_value_cents),
                        "assessed_value": cents_to_dollars(p.assessed_value_cents),
                        "ownership_type": p.ownership_type,
                        "fairness_score": p.fairness_score,
                        "recommended_action": p.recommended_action,
                        "estimated_savings": cents_to_dollars(p.estimated_savings_cents),
                    }
                    for p in properties
                ],
            }

    return SimpleReportGenerator(get_engine())
